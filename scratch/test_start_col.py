import openpyxl
import os
import re

file_path = r"C:\Users\comun\OneDrive\1.Conm Oficina\Ficheros\expedia_revenue_management_1109797_2026_06_10.xlsx"

wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

raw_data = []
for r in range(1, sheet.max_row + 1):
    row = []
    for c in range(1, sheet.max_column + 1):
        row.append(sheet.cell(row=r, column=c).value)
    raw_data.append(row)

# Replicate JS logic to find anchors
guadianaRow = -1
cumbriaRow = -1

for r in range(0, min(len(raw_data), 100)):
    row = raw_data[r] or []
    row_str = " ".join([str(x) for x in row if x is not None]).lower()

    if guadianaRow == -1 and ("guadiana" in row_str or "your property" in row_str) and "tarifas" not in row_str:
        non_empty = [x for x in row if x is not None and x != ""]
        if len(non_empty) > 2:
            guadianaRow = r

    if cumbriaRow == -1 and "cumbria" in row_str:
        cumbriaRow = r

    if guadianaRow != -1 and cumbriaRow != -1:
        break

# Find Date row
dateNumRow = -1
for r in range(guadianaRow - 1, max(-1, guadianaRow - 11), -1):
    row = raw_data[r] or []
    validDayNums = 0
    for c in range(1, min(len(row), 50)):
        cellStr = str(row[c]) if row[c] is not None else ""
        if '%' in cellStr:
            continue
        match = re.search(r'(\d+)', cellStr)
        if match:
            val = int(match.group(1))
            if 1 <= val <= 31:
                validDayNums += 1
    if validDayNums > 3:
        dateNumRow = r
        break

print(f"Date Row: {dateNumRow + 1}")

# Test Approach 1: Expanding the text match in gRowData loop
startCol1 = 1
gRowData = raw_data[guadianaRow]
for c in range(1, len(gRowData)):
    val = gRowData[c]
    if val is not None:
        val_str = str(val).strip().upper()
        is_val = False
        try:
            float(val)
            is_val = True
        except ValueError:
            if val_str in ['S', 'C', 'SOLD OUT', 'SOLD_OUT', 'COMPLETO', 'COMPLETADO', 'CERRADO']:
                is_val = True
        if is_val:
            startCol1 = c
            break

print(f"Approach 1 startCol: {startCol1}")

# Test Approach 2: Running the Date Row double check ALWAYS
startCol2 = 1
# First find by simple loop
for c in range(1, len(gRowData)):
    val = gRowData[c]
    if val is not None:
        try:
            float(val)
            startCol2 = c
            break
        except ValueError:
            if str(val).upper() == 'S':
                startCol2 = c
                break

# Now apply double check without startCol2 == 1 constraint
if dateNumRow != -1 and raw_data[dateNumRow]:
    for c in range(1, len(raw_data[dateNumRow])):
        cellStr = str(raw_data[dateNumRow][c]) if raw_data[dateNumRow][c] is not None else ""
        if '%' in cellStr:
            continue
        match = re.search(r'(\d+)', cellStr)
        if match:
            val = int(match.group(1))
            if 1 <= val <= 31:
                startCol2 = c
                break

print(f"Approach 2 startCol: {startCol2}")
